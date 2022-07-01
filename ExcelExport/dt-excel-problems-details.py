import pycurl
import json
import csv
import certifi
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font
from halo import Halo
import sys
from datetime import datetime

spinner = Halo(text='Loading', spinner='dots')

spinner.start()

spinner.text = 'Initializing ...'

### Setup Variables ###
URL=sys.argv[1]
APITOKEN=sys.argv[2]
DEST_FILENAME=sys.argv[3]
FROM_PARAM=sys.argv[4]

### function to go get the data
def dtApiQuery(endpoint):
	buffer=io.BytesIO()
	c = pycurl.Curl()
	c.setopt(c.URL, URL + endpoint)
	c.setopt(pycurl.CAINFO, certifi.where())
	c.setopt(c.HTTPHEADER, ['Authorization: Api-Token ' + APITOKEN] )
	c.setopt(pycurl.WRITEFUNCTION, buffer.write)
	c.perform()
	spinner.text = endpoint + ' - Status: %d' % c.getinfo(c.RESPONSE_CODE)
	c.close()
	return(buffer.getvalue().decode('UTF-8'))

def toDateTime(timestamp_in_seconds):
	return(datetime.fromtimestamp(int(timestamp_in_seconds)/1000).strftime("%A, %B %d, %Y %I:%M:%S"))

### function to process problems
def processProblems(wsProblemsDetails, nextPageKey, fromParam):

	spinner.text = f'Current Page: {nextPageKey}'

	### Get & Process hosts data
	if nextPageKey is None:
		problemsJson=dtApiQuery('problems?fields=evidenceDetails%2CimpactAnalysis%2CrecentComments&from='+fromParam+'&sort=%2Bstatus%2C-startTime')
		problems=json.loads(problemsJson)
	else:
		problemsJson=dtApiQuery('problems?fields=evidenceDetails%2CimpactAnalysis%2CrecentComments&nextPageKey='+nextPageKey)
		problems=json.loads(problemsJson)

	for problem in problems['problems']:
		problemId = problem['problemId']
		displayId = problem['displayId']
		title = problem['title']
		impactLevel = problem['impactLevel']
		severityLevel = problem['severityLevel']
		status = problem['status']

		affectedEntities = json.dumps(problem['affectedEntities'], indent=4)
		impactedEntities = json.dumps(problem['impactedEntities'], indent=4)
		rootCauseEntity = json.dumps(problem['rootCauseEntity'], indent=4)
		managementZones = json.dumps(problem['managementZones'], indent=4)
		entityTags = json.dumps(problem['entityTags'], indent=4)
		problemFilters = json.dumps(problem['problemFilters'], indent=4)

		startTime = toDateTime(problem['startTime'])
		endTime = toDateTime(problem['endTime'])

		problemsDetailsJson=dtApiQuery('problems/' + problemId)
		problemsDetails=json.loads(problemsDetailsJson)

		problemDetailsProblemId = problemsDetails['problemId']
		problemDetailsDisplayId = problemsDetails['displayId']
		problemDetailsTitle = problemsDetails['title']
		problemDetailsImpactLevel = problemsDetails['impactLevel']
		problemDetailsSeverityLevel = problemsDetails['severityLevel']
		problemDetailsStatus = problemsDetails['status']
		problemDetailsAffectedEntities = json.dumps(problemsDetails['affectedEntities'], indent=4)
		problemDetailsImpactedEntities = json.dumps(problemsDetails['impactedEntities'], indent=4)
		problemDetailsRootCauseEntity = json.dumps(problemsDetails['rootCauseEntity'], indent=4)
		problemDetailsManagementZones = json.dumps(problemsDetails['managementZones'], indent=4)
		problemDetailsEntityTags = json.dumps(problemsDetails['entityTags'], indent=4)
		problemDetailsProblemFilters = json.dumps(problemsDetails['problemFilters'], indent=4)

		problemDetailsStartTime = toDateTime(problemsDetails['startTime'])
		problemDetailsEndTime = toDateTime(problemsDetails['endTime'])

		problemDetailsEvidenceDetails = json.dumps(problemsDetails['evidenceDetails'], indent=4)
		problemDetailsRecentComments = json.dumps(problemsDetails['recentComments'], indent=4)
		problemDetailsImpactAnalysis = json.dumps(problemsDetails['impactAnalysis'], indent=4)

		wsProblemsDetails.append( [ problemId,
			displayId,
			title,
			impactLevel,
			severityLevel,
			status,
			affectedEntities,
			impactedEntities,
			rootCauseEntity,
			managementZones,
			entityTags,
			problemFilters,
			startTime,
			endTime,
			problemDetailsProblemId,
			problemDetailsDisplayId,
			problemDetailsTitle,
			problemDetailsImpactLevel,
			problemDetailsSeverityLevel,
			problemDetailsStatus,
			problemDetailsAffectedEntities,
			problemDetailsImpactedEntities,
			problemDetailsRootCauseEntity,
			problemDetailsManagementZones,
			problemDetailsEntityTags,
			problemDetailsProblemFilters,
			problemDetailsStartTime,
			problemDetailsEndTime,
			problemDetailsEvidenceDetails,
			problemDetailsRecentComments,
			problemDetailsImpactAnalysis 
		] )
	key = problems.get('nextPageKey')
	if key is not None:
		spinner.text = 'Loading new page from problem details ...'
		processProblems(wsProblemsDetails, problems.get('nextPageKey'), fromParam)

### Setup workbook
wb = Workbook()
wsProblemsDetails = wb.create_sheet("Problems Details")
wb.remove(wb.active)

wsProblemsDetails.append( ['problemId','displayId','title','impactLevel','severityLevel','status', 'affectedEntities', 'impactedEntities', 'rootCauseEntity', 'managementZones', 'entityTags', 'problemFilters', 'startTime', 'endTime', 'problemDetailsProblemId', 'problemDetailsDisplayId', 'problemDetailsTitle', 'problemDetailsImpactLevel', 'problemDetailsSeverityLevel', 'problemDetailsStatus', 'problemDetailsAffectedEntities', 'problemDetailsImpactedEntities', 'problemDetailsRootCauseEntity', 'problemDetailsManagementZones', 'problemDetailsEntityTags', 'problemDetailsProblemFilters', 'problemDetailsStartTime', 'problemDetailsEndTime', 'problemDetailsEvidenceDetails', 'problemDetailsRecentComments', 'problemDetailsImpactAnalysis'] )

processProblems(wsProblemsDetails, None, FROM_PARAM)

### set column widths
for ws in wb.worksheets:
	for column_cells in ws.columns:
    		length = max(len(str(cell.value)) for cell in column_cells)

### Set header format
for ws in wb.worksheets:
	for cell in ws["1:1"]:
		cell.style='Headline 3'

wsProblemsDetails.column_dimensions['A'].width = 20
wsProblemsDetails.column_dimensions['B'].width = 20
wsProblemsDetails.column_dimensions['C'].width = 20
wsProblemsDetails.column_dimensions['D'].width = 20
wsProblemsDetails.column_dimensions['E'].width = 20
wsProblemsDetails.column_dimensions['F'].width = 20
wsProblemsDetails.column_dimensions['G'].width = 40
wsProblemsDetails.column_dimensions['H'].width = 40
wsProblemsDetails.column_dimensions['I'].width = 40
wsProblemsDetails.column_dimensions['J'].width = 40
wsProblemsDetails.column_dimensions['K'].width = 40
wsProblemsDetails.column_dimensions['L'].width = 40
wsProblemsDetails.column_dimensions['M'].width = 20
wsProblemsDetails.column_dimensions['N'].width = 20
wsProblemsDetails.column_dimensions['O'].width = 20
wsProblemsDetails.column_dimensions['P'].width = 20
wsProblemsDetails.column_dimensions['Q'].width = 20
wsProblemsDetails.column_dimensions['R'].width = 20
wsProblemsDetails.column_dimensions['S'].width = 20
wsProblemsDetails.column_dimensions['T'].width = 20
wsProblemsDetails.column_dimensions['U'].width = 40
wsProblemsDetails.column_dimensions['V'].width = 40
wsProblemsDetails.column_dimensions['W'].width = 40
wsProblemsDetails.column_dimensions['X'].width = 40
wsProblemsDetails.column_dimensions['Y'].width = 40
wsProblemsDetails.column_dimensions['Z'].width = 40
wsProblemsDetails.column_dimensions['AA'].width = 20
wsProblemsDetails.column_dimensions['AB'].width = 20
wsProblemsDetails.column_dimensions['AC'].width = 40
wsProblemsDetails.column_dimensions['AD'].width = 40
wsProblemsDetails.column_dimensions['AE'].width = 40
wsProblemsDetails.column_dimensions['AF'].width = 40

wsProblemsDetails.auto_filter.ref="A2:AE2"

### Output file
wb.save(filename=DEST_FILENAME)

spinner.succeed('File generated successfully!')

spinner.stop()